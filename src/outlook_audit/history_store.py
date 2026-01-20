import os
from dataclasses import dataclass
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


@dataclass
class HistoryStore:
    history_path: str
    sheet_name: str

    def _final_order(self) -> list[str]:
        return [
            "Código Cliente",
            "Nome do Cliente",
            "Estrutura",
            "Ativo",
            "% PL",
            "Assessor da Operação",
            "Assessor do Cliente",
            "Email Assessor",
            "Email Lider",
            "Status",
            "Data Envio",
            "Assunto",
            "Token Identificador",
            "ConversationID",
            "InternetID",
            "EntryID",
            "Data da Nova Cobrança",
            "Status Auditoria",
            "Data da Resposta",
            "Conteúdo da Resposta",
        ]

    def append_dispatch_record(
        self,
        operation_row: dict,
        email_assessor: str,
        email_lider: str,
        assunto: str,
        token: str,
        status: str,
        conversation_id: str,
        internet_id: str,
        entry_id: str,
    ) -> None:
        df = pd.DataFrame([operation_row])

        df_out = pd.DataFrame()
        df_out["Código Cliente"] = df.get("Código Cliente")
        df_out["Nome do Cliente"] = df.get("Nome do Cliente")
        df_out["Estrutura"] = df.get("Estrutura")
        df_out["Ativo"] = df.get("Ativo")
        df_out["% PL"] = df.get("% PL")
        df_out["Assessor da Operação"] = df.get("Assessor da Operação")
        df_out["Assessor do Cliente"] = df.get("Assessor do Cliente")

        df_out["Email Assessor"] = email_assessor
        df_out["Email Lider"] = email_lider
        df_out["Status"] = status
        df_out["Data Envio"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df_out["Assunto"] = assunto
        df_out["Token Identificador"] = token
        df_out["ConversationID"] = conversation_id
        df_out["InternetID"] = internet_id
        df_out["EntryID"] = entry_id

        df_out["Data da Nova Cobrança"] = None
        df_out["Status Auditoria"] = None
        df_out["Data da Resposta"] = None
        df_out["Conteúdo da Resposta"] = None

        for col in self._final_order():
            if col not in df_out.columns:
                df_out[col] = None
        df_out = df_out[self._final_order()]

        if os.path.exists(self.history_path):
            wb = load_workbook(self.history_path)
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.create_sheet(self.sheet_name)

            if ws.max_row < 1 or all((c.value is None or c.value == "") for c in ws[1]):
                header_df = pd.DataFrame(columns=self._final_order())
                header_row = list(dataframe_to_rows(header_df, index=False, header=True))[0]
                for col_idx, value in enumerate(header_row, start=1):
                    ws.cell(row=1, column=col_idx, value=value)

            last_row = 1
            for r in range(ws.max_row, 1, -1):
                if any(cell.value not in (None, "") for cell in ws[r]):
                    last_row = r
                    break
            next_row = max(last_row + 1, 2)

            rows = list(dataframe_to_rows(df_out, index=False, header=False))
            rows = [r for r in rows if any(v is not None for v in r)]
            for r in rows:
                for col_idx, value in enumerate(r, start=1):
                    ws.cell(row=next_row, column=col_idx, value=value)
                next_row += 1

            wb.save(self.history_path)
            wb.close()
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name

            header_df = pd.DataFrame(columns=self._final_order())
            header_row = list(dataframe_to_rows(header_df, index=False, header=True))[0]
            for col_idx, value in enumerate(header_row, start=1):
                ws.cell(row=1, column=col_idx, value=value)

            rows = list(dataframe_to_rows(df_out, index=False, header=False))
            rows = [r for r in rows if any(v is not None for v in r)]
            row_idx = 2
            for r in rows:
                for col_idx, value in enumerate(r, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1

            wb.save(self.history_path)
            wb.close()

    def load_history_df(self) -> pd.DataFrame:
        return pd.read_excel(self.history_path, sheet_name=self.sheet_name)

    def save_history_df(self, df: pd.DataFrame) -> None:
        wb = load_workbook(self.history_path)
        ws = wb[self.sheet_name]

        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(self.history_path)
        wb.close()
